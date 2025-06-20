"""Data exporter for generating formatted CSV and Excel outputs."""

import pandas as pd
from pathlib import Path
from typing import Optional, Dict, List, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import Settings


class DataExporter:
    """Handles exporting data to CSV and Excel formats with formatting."""
    
    def __init__(self, settings: Optional[Settings] = None):
        self.settings = settings or Settings()
        self.timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
    def export_trades(
        self, 
        df: pd.DataFrame,
        filename: str = "congress_trades",
        include_timestamp: bool = True,
        formats: List[str] = ['csv', 'excel']
    ) -> Dict[str, Path]:
        """
        Export trades data to specified formats.
        
        Args:
            df: DataFrame to export
            filename: Base filename (without extension)
            include_timestamp: Add timestamp to filename
            formats: List of formats to export ('csv', 'excel')
            
        Returns:
            Dictionary of format: filepath
        """
        print(f"\n📤 Exporting data to {', '.join(formats)}...")
        
        exported_files = {}
        
        # Create filename with optional timestamp
        if include_timestamp:
            base_filename = f"{filename}_{self.timestamp}"
        else:
            base_filename = filename
            
        # Export to CSV
        if 'csv' in formats:
            csv_path = self._export_to_csv(df, base_filename)
            exported_files['csv'] = csv_path
            
        # Export to Excel
        if 'excel' in formats:
            excel_path = self._export_to_excel(df, base_filename)
            exported_files['excel'] = excel_path
            
        return exported_files
    
    def export_analysis_package(
        self,
        trades_df: pd.DataFrame,
        patterns: Dict[str, pd.DataFrame],
        filename: str = "nancygate_analysis"
    ) -> Path:
        """
        Export comprehensive analysis package to Excel with multiple sheets.
        
        Args:
            trades_df: Main trades DataFrame
            patterns: Dictionary of pattern DataFrames
            filename: Base filename
            
        Returns:
            Path to exported Excel file
        """
        print(f"\n📊 Creating comprehensive analysis package...")
        
        filepath = self.settings.export_dir / f"{filename}_{self.timestamp}.xlsx"
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Export main trades data
            self._write_trades_sheet(writer, trades_df)
            
            # Export high signal trades
            high_signals = trades_df[trades_df['SignalCategory'].isin(['HIGH', 'VERY_HIGH'])] if 'SignalCategory' in trades_df.columns else pd.DataFrame()
            if not high_signals.empty:
                self._write_sheet(writer, high_signals, 'High_Signal_Trades', highlight_signals=True)
            
            # Export pattern analyses
            for pattern_name, pattern_df in patterns.items():
                if not pattern_df.empty:
                    sheet_name = pattern_name.replace('_', ' ').title()[:31]  # Excel limit
                    self._write_sheet(writer, pattern_df, sheet_name)
            
            # Add summary sheet
            self._write_summary_sheet(writer, trades_df, patterns)
            
        print(f"✅ Analysis package exported to: {filepath}")
        return filepath
    
    def _export_to_csv(self, df: pd.DataFrame, filename: str) -> Path:
        """Export DataFrame to CSV."""
        filepath = self.settings.export_dir / f"{filename}.csv"
        
        # Prepare DataFrame for export
        export_df = self._prepare_export_data(df)
        
        # Export with proper encoding
        export_df.to_csv(filepath, index=False, encoding='utf-8-sig')
        
        print(f"  ✓ CSV exported: {filepath.name}")
        return filepath
    
    def _export_to_excel(self, df: pd.DataFrame, filename: str) -> Path:
        """Export DataFrame to formatted Excel."""
        filepath = self.settings.export_dir / f"{filename}.xlsx"
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            self._write_trades_sheet(writer, df)
            
        print(f"  ✓ Excel exported: {filepath.name}")
        return filepath
    
    def _prepare_export_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """Prepare data for export with formatting."""
        export_df = df.copy()
        
        # Select and order columns
        columns = [
            'Name', 'Party', 'State', 'District',
            'Ticker', 'Company', 'Transaction', 'Amount', 'Trade_Size_USD',
            'Traded', 'Filed', 'DaysToReport',
            'TickerType', 'Description', 'Comments',
            'SignalScore', 'SuspicionScore', 'TotalScore', 'SignalCategory',
            'Signals', 'SignalDetails', 'AdvancedSignals',
            'QuickReport', 'CommitteeAlign', 'UnusualSize', 
            'OptionTrade', 'ClusterTrade', 'ClusterSize'
        ]
        
        # Only include columns that exist in the dataframe
        available_columns = [col for col in columns if col in export_df.columns]
        export_df = export_df[available_columns]
        
        # Format dates
        date_columns = ['Traded', 'Filed', 'ReportDate', 'TransactionDate', 'Date']
        for col in date_columns:
            if col in export_df.columns:
                export_df[col] = pd.to_datetime(export_df[col], errors='coerce').dt.strftime('%Y-%m-%d')
        
        # Format amounts
        amount_columns = ['Amount', 'Trade_Size_USD']
        for col in amount_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].apply(self._format_currency)
        
        # Clean boolean columns
        bool_columns = ['QuickReport', 'CommitteeAlign', 'UnusualSize', 'OptionTrade', 'ClusterTrade']
        for col in bool_columns:
            if col in export_df.columns:
                export_df[col] = export_df[col].fillna(False).infer_objects(copy=False).astype(str)
        
        return export_df
    
    def _format_currency(self, value: Any) -> str:
        """Format currency values."""
        if pd.isna(value) or value == 0:
            return '-'
        
        try:
            value = float(value)
            if value >= 1000000:
                return f"${value/1000000:.1f}M"
            elif value >= 1000:
                return f"${value/1000:.0f}K"
            else:
                return f"${value:.0f}"
        except:
            return str(value)
    
    def _write_trades_sheet(self, writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
        """Write formatted trades sheet to Excel."""
        sheet_name = 'All_Trades'
        
        # Prepare data
        export_df = self._prepare_export_data(df)
        export_df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        # Get worksheet
        worksheet = writer.sheets[sheet_name]
        
        # Apply formatting
        self._format_worksheet(worksheet, export_df, highlight_signals=True)
    
    def _write_sheet(
        self, 
        writer: pd.ExcelWriter, 
        df: pd.DataFrame, 
        sheet_name: str,
        highlight_signals: bool = False
    ) -> None:
        """Write a formatted sheet to Excel."""
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        worksheet = writer.sheets[sheet_name]
        self._format_worksheet(worksheet, df, highlight_signals)
    
    def _format_worksheet(
        self, 
        worksheet, 
        df: pd.DataFrame,
        highlight_signals: bool = False
    ) -> None:
        """Apply formatting to Excel worksheet."""
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-fit columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Conditional formatting for signals
        if highlight_signals and 'SignalCategory' in df.columns:
            # Define fills for different signal levels
            very_high_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            high_fill = PatternFill(start_color="FFA06B", end_color="FFA06B", fill_type="solid")
            medium_fill = PatternFill(start_color="FFD93D", end_color="FFD93D", fill_type="solid")
            
            # Find SignalCategory column
            signal_col_idx = list(df.columns).index('SignalCategory') + 1
            
            # Apply conditional formatting
            for row_idx, row in enumerate(df.itertuples(), start=2):
                signal_category = getattr(row, 'SignalCategory', '')
                
                if signal_category == 'VERY_HIGH':
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = very_high_fill
                elif signal_category == 'HIGH':
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = high_fill
                elif signal_category == 'MEDIUM':
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = medium_fill
        
        # Add table formatting
        # Replace spaces with underscores in table name to avoid Excel error
        table_name = f"Table_{worksheet.title.replace(' ', '_').replace('-', '_')}"
        tab = Table(displayName=table_name, ref=worksheet.dimensions)
        style = TableStyleInfo(
            name="TableStyleMedium2", 
            showFirstColumn=False,
            showLastColumn=False, 
            showRowStripes=True, 
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        worksheet.add_table(tab)
        
        # Freeze top row
        worksheet.freeze_panes = 'A2'
    
    def _write_summary_sheet(
        self, 
        writer: pd.ExcelWriter,
        trades_df: pd.DataFrame,
        patterns: Dict[str, pd.DataFrame]
    ) -> None:
        """Create summary sheet with key metrics and insights."""
        summary_data = []
        
        # Overall statistics
        summary_data.extend([
            ['OVERALL STATISTICS', ''],
            ['Total Trades', len(trades_df)],
            ['Date Range', f"{trades_df['Traded'].min()} to {trades_df['Traded'].max()}" if 'Traded' in trades_df.columns else 'N/A'],
            ['Unique Members', trades_df['Name'].nunique() if 'Name' in trades_df.columns else trades_df['Representative'].nunique() if 'Representative' in trades_df.columns else 'N/A'],
            ['Unique Tickers', trades_df['Ticker'].nunique() if 'Ticker' in trades_df.columns else 'N/A'],
            ['', ''],
            ['SIGNAL SUMMARY', ''],
            ['High Signal Trades', (trades_df['SignalCategory'].isin(['HIGH', 'VERY_HIGH'])).sum() if 'SignalCategory' in trades_df.columns else 0],
            ['Quick Reports (<3 days)', trades_df['QuickReport'].sum() if 'QuickReport' in trades_df.columns else 0],
            ['Options Trades', trades_df['OptionTrade'].sum() if 'OptionTrade' in trades_df.columns else 0],
            ['Cluster Trades', trades_df['ClusterTrade'].sum() if 'ClusterTrade' in trades_df.columns else 0],
            ['', '']
        ])
        
        # Pattern insights
        if patterns:
            summary_data.append(['PATTERN INSIGHTS', ''])
            
            # Top movers from ticker momentum
            if 'ticker_momentum' in patterns and not patterns['ticker_momentum'].empty:
                top_tickers = patterns['ticker_momentum'].head(5)
                summary_data.append(['Top Momentum Tickers', ''])
                for _, ticker in top_tickers.iterrows():
                    summary_data.append([f"  {ticker['Ticker']}", f"Score: {ticker['MomentumScore']}"])
            
            summary_data.append(['', ''])
        
        # Create DataFrame and export
        summary_df = pd.DataFrame(summary_data)
        summary_df.columns = ['Metric', 'Value']
        summary_df.to_excel(writer, sheet_name='Summary', index=False)
        
        # Format summary sheet
        worksheet = writer.sheets['Summary']
        
        # Bold formatting for headers
        for row_idx, (metric, value) in enumerate(summary_df.values, start=2):
            if metric and isinstance(metric, str) and metric.isupper() and value == '':
                cell = worksheet.cell(row=row_idx, column=1)
                cell.font = Font(bold=True, size=12)
                
        # Auto-fit columns
        worksheet.column_dimensions['A'].width = 30
        worksheet.column_dimensions['B'].width = 20
    
    def generate_quick_report(self, df: pd.DataFrame) -> str:
        """Generate a quick text summary of the data."""
        report = []
        report.append("NANCYGATE CONGRESSIONAL TRADING ANALYSIS")
        report.append("=" * 50)
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        
        # Basic stats
        report.append("OVERVIEW:")
        report.append(f"• Total trades: {len(df):,}")
        
        if 'Traded' in df.columns:
            date_range = f"{df['Traded'].min()} to {df['Traded'].max()}"
            report.append(f"• Date range: {date_range}")
        
        member_col = 'Name' if 'Name' in df.columns else 'Representative'
        if member_col in df.columns:
            report.append(f"• Unique members: {df[member_col].nunique()}")
            
        if 'Ticker' in df.columns:
            report.append(f"• Unique tickers: {df['Ticker'].nunique()}")
        
        # Signal summary
        if 'SignalCategory' in df.columns:
            report.append("")
            report.append("SIGNAL ANALYSIS:")
            high_signals = (df['SignalCategory'].isin(['HIGH', 'VERY_HIGH'])).sum()
            report.append(f"• High signal trades: {high_signals} ({high_signals/len(df)*100:.1f}%)")
        
        # Top trades
        if 'SignalScore' in df.columns and 'Ticker' in df.columns:
            report.append("")
            report.append("TOP SIGNAL TRADES:")
            top_trades = df.nlargest(5, 'SignalScore')[['Ticker', member_col, 'SignalScore', 'Signals']]
            for _, trade in top_trades.iterrows():
                report.append(f"• {trade['Ticker']} by {trade[member_col]} (Score: {trade['SignalScore']})")
        
        return "\n".join(report) 